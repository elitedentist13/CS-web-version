"""Convert Clinic Solution CSV extracts into a multi-sheet .xlsx"""
import csv
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font


def load_csv(path: Path):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    return rows


def write_sheet(wb, title, rows):
    ws = wb.create_sheet(title)
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    return ws


def main():
    if len(sys.argv) < 4:
        print("Usage: csv-to-xlsx.py notes.csv extend.csv patients.csv [out.xlsx]")
        sys.exit(2)
    notes, extend, patients = map(Path, sys.argv[1:4])
    out = Path(sys.argv[4]) if len(sys.argv) > 4 else notes.with_name(
        notes.name.replace("_notes.csv", ".xlsx")
    )

    wb = Workbook()
    # remove default sheet
    default = wb.active
    wb.remove(default)

    write_sheet(wb, "ConsultationNotes", load_csv(notes))
    write_sheet(wb, "ExtendMedicalData", load_csv(extend))
    write_sheet(wb, "Patients", load_csv(patients))
    wb.save(out)
    print(f"XLSX_OK {out}")


if __name__ == "__main__":
    main()
